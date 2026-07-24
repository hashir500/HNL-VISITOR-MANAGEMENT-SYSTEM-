from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.urls import reverse
from openpyxl import load_workbook
from .models import visitor_card, visitor, company_type
import os
from urllib.parse import urlencode
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.conf import settings


@login_required
@permission_required("visitors.view_visitor_card", raise_exception=True)
def visitor_card_list(request):
    cards = visitor_card.objects.all().order_by("id")

    return render(request, "visitors/visitor_card_list.html", {
        "cards": cards,
    })


@login_required
@permission_required("visitors.add_visitor_card", raise_exception=True)
def visitor_card_create(request):
    if request.method == "POST":
        crd_no = request.POST.get("CRD_No", "").strip()
        crd_desc = request.POST.get("CRD_Desc", "").strip()
        crd_active = True if request.POST.get("CRD_Active") else False

        if visitor_card.objects.filter(CRD_No__iexact=crd_no).exists():
            messages.error(request, f"Card number {crd_no} already exists.")
            return redirect("visitor_card_list")

        visitor_card.objects.create(
            CRD_No=crd_no,
            CRD_Desc=crd_desc,
            CRD_Active=crd_active,
        )

        messages.success(request, "Visitor card created successfully.")

    return redirect("visitor_card_list")


@login_required
@permission_required("visitors.change_visitor_card", raise_exception=True)
def visitor_card_update(request, pk):
    card = get_object_or_404(visitor_card, pk=pk)

    if request.method == "POST":
        crd_no = request.POST.get("CRD_No", "").strip()
        crd_desc = request.POST.get("CRD_Desc", "").strip()
        crd_active = True if request.POST.get("CRD_Active") else False

        if visitor_card.objects.filter(CRD_No__iexact=crd_no).exclude(pk=pk).exists():
            messages.error(request, f"Card number {crd_no} already exists.")
            return redirect("visitor_card_list")

        card.CRD_No = crd_no
        card.CRD_Desc = crd_desc
        card.CRD_Active = crd_active
        card.save()

        messages.success(request, "Visitor card updated successfully.")

    return redirect("visitor_card_list")


@login_required
@permission_required("visitors.delete_visitor_card", raise_exception=True)
def visitor_card_delete(request, pk):
    card = get_object_or_404(visitor_card, pk=pk)

    if request.method == "POST":
        card.delete()
        messages.success(request, "Visitor card deleted successfully.")

    return redirect("visitor_card_list")


# Company Type Views

@login_required
@permission_required("visitors.view_company_type", raise_exception=True)
def company_type_list(request):
    search = (request.GET.get("search") or "").strip()
    company_types = company_type.objects.all().order_by("-created_at")

    if search:
        company_types = company_types.filter(company_name__icontains=search)

    return render(
        request,
        "visitors/company_type_list.html",
        {
            "company_types": company_types,
            "search": search,
        },
    )


@login_required
@permission_required("visitors.add_company_type", raise_exception=True)
def company_type_create(request):
    if request.method == "POST":
        company_name = (request.POST.get("company_name") or "").strip()

        if not company_name:
            messages.error(request, "Company type is required.")
            return redirect("company_type_list")

        if company_type.objects.filter(company_name__iexact=company_name).exists():
            messages.error(request, f"Company type '{company_name}' already exists.")
            return redirect("company_type_list")

        company_type.objects.create(company_name=company_name)
        messages.success(request, "Company type created successfully.")

    return redirect("company_type_list")


@login_required
@permission_required("visitors.change_company_type", raise_exception=True)
def company_type_update(request, pk):
    comp_type = get_object_or_404(company_type, pk=pk)

    if request.method == "POST":
        company_name = (request.POST.get("company_name") or "").strip()

        if not company_name:
            messages.error(request, "Company type is required.")
            return redirect("company_type_list")

        if company_type.objects.filter(company_name__iexact=company_name).exclude(pk=pk).exists():
            messages.error(request, f"Company type '{company_name}' already exists.")
            return redirect("company_type_list")

        comp_type.company_name = company_name
        comp_type.save(update_fields=["company_name"])

        messages.success(request, "Company type updated successfully.")

    return redirect("company_type_list")


@login_required
@permission_required("visitors.delete_company_type", raise_exception=True)
def company_type_delete(request, pk):
    comp_type = get_object_or_404(company_type, pk=pk)

    if request.method == "POST":
        company_name = comp_type.company_name
        try:
            comp_type.delete()
            messages.success(request, f"Company type '{company_name}' deleted successfully.")
        except ProtectedError:
            messages.error(request, f"Company type '{company_name}' cannot be deleted because it is in use.")

    return redirect("company_type_list")


# visitor views

@login_required
@permission_required("visitors.view_visitor", raise_exception=True)
def visitor_list(request):
    search = (request.GET.get("search") or "").strip()

    visitors = visitor.objects.select_related("company_type").all().order_by("-visitor_created_at")
    company_types = company_type.objects.all().order_by("company_name")

    if search:
        visitors = visitors.filter(
            Q(visitor_name__icontains=search)
            | Q(visitor_phone__icontains=search)
            | Q(visitor_cnic__icontains=search)
            | Q(visitor_address__icontains=search)
            | Q(visitor_company__icontains=search)
            | Q(company_type__company_name__icontains=search)
        )

    return render(
        request,
        "visitors/visitor_list.html",
        {
            "visitors": visitors,
            "company_types": company_types,
            "search": search,
            "open_add_visitor": request.GET.get("open_add_visitor"),
            "next_page": request.GET.get("next", ""),
            "form_name": request.GET.get("visitor_name", ""),
            "form_phone": request.GET.get("visitor_phone", ""),
            "form_cnic": request.GET.get("visitor_cnic", ""),
            "form_company": request.GET.get("visitor_company", ""),
            "form_company_type": request.GET.get("company_type", ""),
            "form_address": request.GET.get("visitor_address", ""),
        },
    )


@login_required
@permission_required("visitors.add_visitor", raise_exception=True)
def visitor_create(request):
    if request.method != "POST":
        return redirect("visitor_list")

    visitor_name = (request.POST.get("visitor_name") or "").strip()
    visitor_phone = (request.POST.get("visitor_phone") or "").strip()
    visitor_cnic = (request.POST.get("visitor_cnic") or "").strip()
    visitor_address = (request.POST.get("visitor_address") or "").strip()
    visitor_company = (request.POST.get("visitor_company") or "").strip()
    company_type_id = request.POST.get("company_type") or None

    next_page = (request.POST.get("next") or "").strip()

    if not visitor_name:
        messages.error(request, "Visitor name is required.")
        return _visitor_form_redirect(
            next_page, visitor_name, visitor_phone, visitor_cnic, visitor_company, company_type_id, visitor_address
        )

    if not visitor_phone:
        messages.error(request, "Visitor phone is required.")
        return _visitor_form_redirect(
            next_page, visitor_name, visitor_phone, visitor_cnic, visitor_company, company_type_id, visitor_address
        )

    normalized_cnic = normalize_cnic(visitor_cnic) if visitor_cnic else None

    if normalized_cnic and visitor.objects.filter(visitor_cnic=normalized_cnic).exists():
        messages.error(request, f"A visitor with CNIC {normalized_cnic} already exists.")
        return _visitor_form_redirect(
            next_page, visitor_name, visitor_phone, visitor_cnic, visitor_company, company_type_id, visitor_address
        )

    comp_type_obj = None
    if company_type_id:
        comp_type_obj = company_type.objects.filter(pk=company_type_id).first()

    new_visitor = visitor.objects.create(
        visitor_name=visitor_name,
        visitor_phone=visitor_phone,
        visitor_cnic=normalized_cnic,
        visitor_address=visitor_address,
        visitor_company=visitor_company,
        company_type=comp_type_obj,
    )

    messages.success(request, "Visitor created successfully.")

    if next_page == "visits":
        query_string = urlencode({
            "open_add_visit": "1",
            "visitor_id": new_visitor.visitor_id,
        })
        return redirect(f"{reverse('visit_list')}?{query_string}")

    if next_page == "backlogs":
        query_string = urlencode({
            "open_add_backlog": "1",
            "selected_visitor_id": new_visitor.visitor_id,
        })
        return redirect(f"{reverse('backlog_list_create')}?{query_string}")

    return redirect("visitor_list")


@login_required
@permission_required("visitors.change_visitor", raise_exception=True)
def visitor_update(request, visitor_id):
    visitor_obj = get_object_or_404(visitor, visitor_id=visitor_id)

    if request.method != "POST":
        return redirect("visitor_list")

    visitor_name = (request.POST.get("visitor_name") or "").strip()
    visitor_phone = (request.POST.get("visitor_phone") or "").strip()
    visitor_cnic = (request.POST.get("visitor_cnic") or "").strip()
    visitor_address = (request.POST.get("visitor_address") or "").strip()
    visitor_company = (request.POST.get("visitor_company") or "").strip()
    company_type_id = request.POST.get("company_type") or None

    if not visitor_name:
        messages.error(request, "Visitor name is required.")
        return redirect("visitor_list")

    if not visitor_phone:
        messages.error(request, "Visitor phone is required.")
        return redirect("visitor_list")

    normalized_cnic = normalize_cnic(visitor_cnic) if visitor_cnic else None

    if normalized_cnic:
        duplicate_cnic = (
            visitor.objects
            .filter(visitor_cnic=normalized_cnic)
            .exclude(visitor_id=visitor_obj.visitor_id)
            .exists()
        )

        if duplicate_cnic:
            messages.error(request, f"Another visitor with CNIC {normalized_cnic} already exists.")
            return redirect("visitor_list")

    comp_type_obj = None
    if company_type_id:
        comp_type_obj = company_type.objects.filter(pk=company_type_id).first()

    visitor_obj.visitor_name = visitor_name
    visitor_obj.visitor_phone = visitor_phone
    visitor_obj.visitor_cnic = normalized_cnic
    visitor_obj.visitor_address = visitor_address
    visitor_obj.visitor_company = visitor_company
    visitor_obj.company_type = comp_type_obj

    visitor_obj.save(
        update_fields=[
            "visitor_name",
            "visitor_phone",
            "visitor_cnic",
            "visitor_address",
            "visitor_company",
            "company_type",
        ]
    )

    messages.success(request, "Visitor updated successfully.")
    return redirect("visitor_list")


@login_required
@permission_required("visitors.delete_visitor", raise_exception=True)
def visitor_delete(request, visitor_id):
    visitor_obj = get_object_or_404(visitor, visitor_id=visitor_id)

    if request.method == "POST":
        visitor_name = visitor_obj.visitor_name

        try:
            visitor_obj.delete()
            messages.success(request, f"{visitor_name} deleted successfully.")
        except ProtectedError:
            messages.error(request, "This visitor cannot be deleted because visit records are linked to them.")

    return redirect("visitor_list")


def normalize_cnic(cnic):
    return "".join(character for character in str(cnic) if character.isdigit())


def _visitor_form_redirect(next_page, name="", phone="", cnic="", company="", comp_type="", address=""):
    """
    Reopen the Add Visitor modal and preserve inputs when validation fails.
    """
    query_parameters = {"open_add_visitor": "1"}

    if next_page:
        query_parameters["next"] = next_page
    if name:
        query_parameters["visitor_name"] = name
    if phone:
        query_parameters["visitor_phone"] = phone
    if cnic:
        query_parameters["visitor_cnic"] = cnic
    if company:
        query_parameters["visitor_company"] = company
    if comp_type:
        query_parameters["company_type"] = comp_type
    if address:
        query_parameters["visitor_address"] = address

    return redirect(f"{reverse('visitor_list')}?{urlencode(query_parameters)}")


# visitor card import
@login_required
@permission_required("visitors.add_visitor_card", raise_exception=True)
def visitor_card_import_upload(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("visitor_card_list")

        upload_dir = os.path.join(settings.MEDIA_ROOT, "imports")
        os.makedirs(upload_dir, exist_ok=True)

        file_path = os.path.join(upload_dir, excel_file.name)

        with open(file_path, "wb+") as destination:
            for chunk in excel_file.chunks():
                destination.write(chunk)

        workbook = load_workbook(file_path)
        sheet = workbook.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

        request.session["visitor_card_import_file"] = file_path
        request.session["visitor_card_import_headers"] = headers

        return render(request, "visitors/visitor_card_import_mapping.html", {
            "headers": headers,
        })

    return redirect("visitor_card_list")


@login_required
@permission_required("visitors.add_visitor_card", raise_exception=True)
def visitor_card_import_process(request):
    if request.method == "POST":
        file_path = request.session.get("visitor_card_import_file")

        if not file_path or not os.path.exists(file_path):
            messages.error(request, "Import file not found. Please upload again.")
            return redirect("visitor_card_list")

        workbook = load_workbook(file_path)
        sheet = workbook.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

        mapping = {
            "CRD_No": request.POST.get("CRD_No"),
            "CRD_Desc": request.POST.get("CRD_Desc"),
            "CRD_Active": request.POST.get("CRD_Active"),
        }

        def get_value(row, field):
            excel_col = mapping.get(field)

            if not excel_col or excel_col not in headers:
                return None

            col_index = headers.index(excel_col) + 1
            value = sheet.cell(row=row, column=col_index).value

            if value is None:
                return None

            if isinstance(value, (int, float)):
                if float(value).is_integer():
                    return str(int(value))
                return str(value)

            return str(value).strip()

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row in range(2, sheet.max_row + 1):
            crd_no = get_value(row, "CRD_No")
            crd_desc = get_value(row, "CRD_Desc")
            crd_active = get_value(row, "CRD_Active")

            if not crd_no:
                skipped_count += 1
                continue

            active_text = str(crd_active).strip().lower() if crd_active else "true"
            is_active = active_text in ["true", "1", "yes", "y", "active"]

            card, created = visitor_card.objects.update_or_create(
                CRD_No=crd_no,
                defaults={
                    "CRD_Desc": crd_desc or "",
                    "CRD_Active": is_active,
                }
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        request.session.pop("visitor_card_import_file", None)
        request.session.pop("visitor_card_import_headers", None)

        messages.success(
            request,
            f"Import completed. Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_count}."
        )

    return redirect("visitor_card_list")


@login_required
@permission_required("visitors.delete_visitor_card", raise_exception=True)
def visitor_card_delete_all(request):
    if request.method == "POST":
        visitor_card.objects.all().delete()
        messages.success(request, "All visitor cards deleted successfully.")

    return redirect("visitor_card_list")