from urllib.parse import urlencode
from django.shortcuts import render, redirect, get_object_or_404
import requests
import openpyxl
from collections import defaultdict

from django.contrib import messages
from django.conf import settings
import os
import uuid
from openpyxl import load_workbook
from django.contrib.auth.models import Group, Permission
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.db.models import Q
from django.contrib.auth.decorators import user_passes_test
import MySQLdb

from django.contrib.auth.decorators import login_required, permission_required
from .access import employees_visible_to_user
from .models import sys_usr_system
from .models import sys_cmp_master
from .forms import CompanyMasterForm
from .models import sys_bra_master
from .forms import BranchMasterForm
from .models import sys_div_master
from .forms import DivisionMasterForm
from .models import sys_dep_master, sys_div_master
from .forms import DepartmentMasterForm
from .models import sys_emp_master
from .forms import EmployeeMasterForm
from .models import sys_pur_master
from .forms import PurposeMasterForm
from django.http import JsonResponse


@login_required
@permission_required("masters.view_sys_cmp_master", raise_exception=True)
def company_list(request):
    companies = sys_cmp_master.objects.all().order_by("id")
    form = CompanyMasterForm()

    return render(request, "masters/company_list.html", {
        "companies": companies,
        "form": form,
    })


@login_required
@permission_required("masters.add_sys_cmp_master", raise_exception=True)
def company_create(request):
    if request.method == "POST":
        form = CompanyMasterForm(request.POST)
        if form.is_valid():
            form.save()

    return redirect("company_list")


@login_required
@permission_required("masters.change_sys_cmp_master", raise_exception=True)
def company_update(request, pk):
    company = get_object_or_404(sys_cmp_master, pk=pk)

    if request.method == "POST":
        form = CompanyMasterForm(request.POST, instance=company)
        if form.is_valid():
            form.save()

    return redirect("company_list")


@login_required
@permission_required("masters.delete_sys_cmp_master", raise_exception=True)
def company_delete(request, pk):
    company = get_object_or_404(sys_cmp_master, pk=pk)

    if request.method == "POST":
        company.delete()

    return redirect("company_list")


@login_required
@permission_required("masters.view_sys_bra_master", raise_exception=True)
def branch_list(request):
    branches = sys_bra_master.objects.all().order_by("bra_code")

    return render(request, "masters/branch_list.html", {
        "branches": branches,
    })


@login_required
@permission_required("masters.add_sys_bra_master", raise_exception=True)
def branch_create(request):
    if request.method == "POST":
        form = BranchMasterForm(request.POST)
        if form.is_valid():
            form.save()

    return redirect("branch_list")


@login_required
@permission_required("masters.change_sys_bra_master", raise_exception=True)
def branch_update(request, pk):
    branch = get_object_or_404(sys_bra_master, pk=pk)

    if request.method == "POST":
        form = BranchMasterForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()

    return redirect("branch_list")


@login_required
@permission_required("masters.delete_sys_bra_master", raise_exception=True)
def branch_delete(request, pk):
    branch = get_object_or_404(sys_bra_master, pk=pk)

    if request.method == "POST":
        branch.delete()

    return redirect("branch_list")


@login_required
@permission_required("masters.view_sys_div_master", raise_exception=True)
def division_list(request):
    divisions = sys_div_master.objects.all().order_by("div_code")

    return render(request, "masters/division_list.html", {
        "divisions": divisions,
    })


@login_required
@permission_required("masters.add_sys_div_master", raise_exception=True)
def division_create(request):
    if request.method == "POST":
        form = DivisionMasterForm(request.POST)
        if form.is_valid():
            form.save()

    return redirect("division_list")


@login_required
@permission_required("masters.change_sys_div_master", raise_exception=True)
def division_update(request, pk):
    division = get_object_or_404(sys_div_master, pk=pk)

    if request.method == "POST":
        form = DivisionMasterForm(request.POST, instance=division)
        if form.is_valid():
            form.save()

    return redirect("division_list")


@login_required
@permission_required("masters.delete_sys_div_master", raise_exception=True)
def division_delete(request, pk):
    division = get_object_or_404(sys_div_master, pk=pk)

    if request.method == "POST":
        division.delete()

    return redirect("division_list")


@login_required
@permission_required("masters.view_sys_dep_master", raise_exception=True)
def department_master_list(request):
    departments = (
        sys_dep_master.objects
        .select_related("dep_div_code")
        .all()
        .order_by("dep_code")
    )

    divisions = sys_div_master.objects.filter(div_active=True).order_by("div_desc")

    return render(request, "masters/department_list.html", {
        "departments": departments,
        "divisions": divisions,
    })


@login_required
@permission_required("masters.add_sys_dep_master", raise_exception=True)
def department_master_create(request):
    if request.method == "POST":
        form = DepartmentMasterForm(request.POST)
        if form.is_valid():
            form.save()

    return redirect("department_master_list")


@login_required
@permission_required("masters.change_sys_dep_master", raise_exception=True)
def department_master_update(request, pk):
    department = get_object_or_404(sys_dep_master, pk=pk)

    if request.method == "POST":
        form = DepartmentMasterForm(request.POST, instance=department)
        if form.is_valid():
            form.save()

    return redirect("department_master_list")


@login_required
@permission_required("masters.delete_sys_dep_master", raise_exception=True)
def department_master_delete(request, pk):
    department = get_object_or_404(sys_dep_master, pk=pk)

    if request.method == "POST":
        department.delete()

    return redirect("department_master_list")


@login_required
@permission_required("masters.add_sys_dep_master", raise_exception=True)
def department_import_upload(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("department_master_list")

        upload_dir = os.path.join(settings.MEDIA_ROOT, "imports")
        os.makedirs(upload_dir, exist_ok=True)

        file_name = f"{uuid.uuid4()}_{excel_file.name}"
        file_path = os.path.join(upload_dir, file_name)

        with open(file_path, "wb+") as destination:
            for chunk in excel_file.chunks():
                destination.write(chunk)

        workbook = load_workbook(file_path)
        sheet = workbook.active

        excel_columns = []
        for cell in sheet[1]:
            if cell.value:
                excel_columns.append(str(cell.value).strip())

        request.session["department_import_file"] = file_path

        return render(request, "masters/department_import_map.html", {
            "excel_columns": excel_columns,
            "db_fields": [
                ("dep_code", "Department Code"),
                ("dep_desc", "Department Description"),
                ("dep_div_code", "Division Code"),
                ("dep_active", "Active"),
            ],
        })

    return redirect("department_master_list")


@login_required
@permission_required("masters.add_sys_dep_master", raise_exception=True)
def department_import_process(request):
    if request.method == "POST":
        file_path = request.session.get("department_import_file")

        if not file_path or not os.path.exists(file_path):
            messages.error(request, "Import file not found. Please upload again.")
            return redirect("department_master_list")

        workbook = load_workbook(file_path)
        sheet = workbook.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

        mapping = {
            "dep_code": request.POST.get("dep_code"),
            "dep_desc": request.POST.get("dep_desc"),
            "dep_div_code": request.POST.get("dep_div_code"),
            "dep_active": request.POST.get("dep_active"),
        }

        imported = 0
        updated = 0
        skipped = 0

        for row in range(2, sheet.max_row + 1):
            try:
                def get_value(db_field):
                    excel_col = mapping.get(db_field)
                    if not excel_col or excel_col not in headers:
                        return None

                    col_index = headers.index(excel_col) + 1
                    value = sheet.cell(row=row, column=col_index).value

                    if value is None:
                        return None

                    return str(value).strip()

                dep_code = get_value("dep_code")
                dep_desc = get_value("dep_desc")
                div_code = get_value("dep_div_code")
                dep_active_value = get_value("dep_active")

                if not dep_code or not dep_desc or not div_code:
                    skipped += 1
                    continue

                division = sys_div_master.objects.filter(div_code=div_code).first()

                if not division:
                    skipped += 1
                    continue

                dep_active = True
                if dep_active_value:
                    dep_active = dep_active_value.lower() in ["true", "1", "yes", "y", "active"]

                obj, created = sys_dep_master.objects.update_or_create(
                    dep_code=dep_code,
                    defaults={
                        "dep_desc": dep_desc,
                        "dep_div_code": division,
                        "dep_active": dep_active,
                    }
                )

                if created:
                    imported += 1
                else:
                    updated += 1

            except Exception:
                skipped += 1

        messages.success(
            request,
            f"Department import completed. Imported: {imported}, Updated: {updated}, Skipped: {skipped}"
        )

        return redirect("department_master_list")

    return redirect("department_master_list")


@login_required
@permission_required("masters.delete_sys_dep_master", raise_exception=True)
def department_delete_all(request):
    if request.method == "POST":
        deleted = sys_dep_master.objects.count()
        sys_dep_master.objects.all().delete()

        messages.success(
            request,
            f"{deleted} departments deleted successfully."
        )

    return redirect("department_master_list")


@login_required
@permission_required("masters.view_sys_emp_master", raise_exception=True)
def employee_master_list(request):
    search = (request.GET.get("search") or "").strip()

    employees = (
        employees_visible_to_user(request.user)
        .select_related(
            "emp_cmp",
            "emp_bra_code",
            "emp_dep_code",
        )
        .order_by("emp_pno")
    )

    if search:
        employees = employees.filter(
            Q(emp_pno__icontains=search)
            | Q(emp_name__icontains=search)
            | Q(emp_designation__icontains=search)
            | Q(emp_email__icontains=search)
            | Q(emp_mobile__icontains=search)
            | Q(emp_phone__icontains=search)
            | Q(emp_cmp__cmp_desc__icontains=search)
            | Q(emp_bra_code__bra_desc__icontains=search)
            | Q(emp_dep_code__dep_desc__icontains=search)
        )

    companies = sys_cmp_master.objects.all().order_by("cmp_code")
    branches = sys_bra_master.objects.all().order_by("bra_code")
    departments = sys_dep_master.objects.all().order_by("dep_code")

    return render(request, "masters/employee_list.html", {
        "employees": employees,
        "companies": companies,
        "branches": branches,
        "departments": departments,
        "search": search,
    })


@login_required
@permission_required("masters.add_sys_emp_master", raise_exception=True)
def employee_master_create(request):
    if request.method == "POST":
        form = EmployeeMasterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Employee added successfully.")
        else:
            messages.error(request, "Employee could not be added.")

    return redirect("employee_master_list")


@login_required
@permission_required("masters.change_sys_emp_master", raise_exception=True)
def employee_master_update(request, pk):
    employee = get_object_or_404(
        employees_visible_to_user(request.user),
        pk=pk,
    )

    if request.method == "POST":
        form = EmployeeMasterForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, "Employee updated successfully.")
        else:
            messages.error(request, "Employee could not be updated.")

    return redirect("employee_master_list")


@login_required
@permission_required("masters.delete_sys_emp_master", raise_exception=True)
def employee_master_delete(request, pk):
    employee = get_object_or_404(
        employees_visible_to_user(request.user),
        pk=pk,
    )

    if request.method == "POST":
        employee.delete()
        messages.success(request, "Employee deleted successfully.")

    return redirect("employee_master_list")


@login_required
@permission_required("masters.add_sys_emp_master", raise_exception=True)
def employee_import_upload(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("employee_master_list")

        upload_dir = os.path.join(settings.MEDIA_ROOT, "imports")
        os.makedirs(upload_dir, exist_ok=True)

        file_name = f"{uuid.uuid4()}_{excel_file.name}"
        file_path = os.path.join(upload_dir, file_name)

        with open(file_path, "wb+") as destination:
            for chunk in excel_file.chunks():
                destination.write(chunk)

        workbook = load_workbook(file_path)
        sheet = workbook.active

        excel_columns = [
            str(cell.value).strip()
            for cell in sheet[1]
            if cell.value
        ]

        request.session["employee_import_file"] = file_path

        return render(request, "masters/employee_import_map.html", {
            "excel_columns": excel_columns,
            "db_fields": [
                ("emp_cmp", "Company Code"),
                ("emp_bra_code", "Branch Code"),
                ("emp_pno", "Employee PNO"),
                ("emp_name", "Employee Name"),
                ("emp_designation", "Designation"),
                ("emp_dep_code", "Department Code"),
                ("emp_email", "Email"),
                ("emp_mobile", "Mobile"),
                ("emp_phone", "Phone"),
                ("emp_pbx", "PBX"),
                ("emp_active", "Active"),
            ],
        })

    return redirect("employee_master_list")


@login_required
@permission_required("masters.add_sys_emp_master", raise_exception=True)
def employee_import_process(request):
    if request.method == "POST":
        file_path = request.session.get("employee_import_file")

        if not file_path or not os.path.exists(file_path):
            messages.error(request, "Import file not found. Please upload again.")
            return redirect("employee_master_list")

        workbook = load_workbook(file_path)
        sheet = workbook.active

        headers = [
            str(cell.value).strip() if cell.value else ""
            for cell in sheet[1]
        ]

        mapping = {
            "emp_cmp": request.POST.get("emp_cmp"),
            "emp_bra_code": request.POST.get("emp_bra_code"),
            "emp_pno": request.POST.get("emp_pno"),
            "emp_name": request.POST.get("emp_name"),
            "emp_designation": request.POST.get("emp_designation"),
            "emp_dep_code": request.POST.get("emp_dep_code"),
            "emp_email": request.POST.get("emp_email"),
            "emp_mobile": request.POST.get("emp_mobile"),
            "emp_phone": request.POST.get("emp_phone"),
            "emp_pbx": request.POST.get("emp_pbx"),
            "emp_active": request.POST.get("emp_active"),
        }

        imported = 0
        updated = 0
        skipped = 0

        def get_value(row, db_field):
            excel_col = mapping.get(db_field)

            if not excel_col or excel_col not in headers:
                return None

            col_index = headers.index(excel_col) + 1
            value = sheet.cell(row=row, column=col_index).value

            if value is None:
                return None
            
            if isinstance(value, float) and value.is_integer():
                value = int(value)

            return str(value).strip()

        for row in range(2, sheet.max_row + 1):
            try:
                cmp_code = get_value(row, "emp_cmp")
                bra_code = get_value(row, "emp_bra_code")
                emp_pno = get_value(row, "emp_pno")
                emp_name = get_value(row, "emp_name")
                emp_designation = get_value(row, "emp_designation")
                dep_code = get_value(row, "emp_dep_code")

                if not cmp_code or not bra_code or not emp_pno or not emp_name or not dep_code:
                    skipped += 1
                    continue

                company = sys_cmp_master.objects.filter(cmp_code=cmp_code).first()
                branch = sys_bra_master.objects.filter(bra_code=bra_code).first()
                department = sys_dep_master.objects.filter(dep_code=dep_code).first()

                if not company or not branch or not department:
                    skipped += 1
                    continue

                emp_active_value = get_value(row, "emp_active")
                emp_active = True

                if emp_active_value:
                    emp_active = emp_active_value.lower() in ["true", "1", "yes", "y", "active"]

                obj, created = sys_emp_master.objects.update_or_create(
                    emp_pno=emp_pno,
                    defaults={
                        "emp_cmp": company,
                        "emp_bra_code": branch,
                        "emp_name": emp_name,
                        "emp_designation": emp_designation or "",
                        "emp_dep_code": department,
                        "emp_email": get_value(row, "emp_email"),
                        "emp_mobile": get_value(row, "emp_mobile"),
                        "emp_phone": get_value(row, "emp_phone"),
                        "emp_pbx": get_value(row, "emp_pbx"),
                        "emp_active": emp_active,
                    }
                )

                if created:
                    imported += 1
                else:
                    updated += 1

            except Exception:
                skipped += 1

        messages.success(
            request,
            f"Employee import completed. Imported: {imported}, Updated: {updated}, Skipped: {skipped}"
        )

        return redirect("employee_master_list")

    return redirect("employee_master_list")


@login_required
@permission_required("masters.delete_sys_emp_master", raise_exception=True)
def employee_delete_all(request):
    if request.method == "POST":
        deleted = sys_emp_master.objects.count()

        sys_emp_master.objects.all().delete()

        messages.success(
            request,
            f"{deleted} employees deleted successfully."
        )

    return redirect("employee_master_list")


@login_required
@permission_required("masters.view_sys_pur_master", raise_exception=True)
def purpose_list(request):
    purposes = sys_pur_master.objects.all().order_by("pur_id")

    return render(request, "masters/purpose_list.html", {
        "purposes": purposes,
    })


@login_required
@permission_required("masters.add_sys_pur_master", raise_exception=True)
def purpose_create(request):
    if request.method == "POST":
        form = PurposeMasterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Purpose added successfully.")
        else:
            messages.error(request, "Purpose could not be added.")

    return redirect("purpose_list")


@login_required
@permission_required("masters.change_sys_pur_master", raise_exception=True)
def purpose_update(request, pk):
    purpose = get_object_or_404(sys_pur_master, pk=pk)

    if request.method == "POST":
        form = PurposeMasterForm(request.POST, instance=purpose)
        if form.is_valid():
            form.save()
            messages.success(request, "Purpose updated successfully.")
        else:
            messages.error(request, "Purpose could not be updated.")

    return redirect("purpose_list")


@login_required
@permission_required("masters.delete_sys_pur_master", raise_exception=True)
def purpose_delete(request, pk):
    purpose = get_object_or_404(sys_pur_master, pk=pk)

    if request.method == "POST":
        purpose.delete()
        messages.success(request, "Purpose deleted successfully.")

    return redirect("purpose_list")


@login_required
@permission_required(
    "masters.view_sys_emp_master",
    raise_exception=True,
)
def fetch_employee_details(request, emp_pno):
    employee = (
        sys_emp_master.objects
        .select_related(
            "emp_dep_code",
            "emp_bra_code",
            "emp_cmp",
        )
        .filter(emp_pno=emp_pno)
        .first()
    )

    if not employee:
        return JsonResponse({
            "success": False,
            "message": "Employee not found.",
        })

    full_name = (employee.emp_name or "").strip()
    name_parts = full_name.split(" ", 1)

    first_name = name_parts[0] if name_parts else ""
    last_name = name_parts[1] if len(name_parts) > 1 else ""

    return JsonResponse({
        "success": True,
        "first_name": first_name,
        "last_name": last_name,
        "designation": employee.emp_designation or "",
        "department": (
            employee.emp_dep_code.dep_code
            if employee.emp_dep_code
            else ""
        ),
        "branch": (
            employee.emp_bra_code.bra_code
            if employee.emp_bra_code
            else "ALL"
        ),
        "company": (
            employee.emp_cmp.cmp_code
            if employee.emp_cmp
            else "ALL"
        ),
        "mobile": employee.emp_mobile or "",
        "email": employee.emp_email or "",
        "phone": employee.emp_phone or "",
    })


def get_optional_company(company_code):
    company_code = (company_code or "").strip()

    if not company_code or company_code.upper() == "ALL":
        return None

    return get_object_or_404(
        sys_cmp_master,
        cmp_code=company_code,
    )


def get_optional_branch(branch_code):
    branch_code = (branch_code or "").strip()

    if not branch_code or branch_code.upper() == "ALL":
        return None

    return get_object_or_404(
        sys_bra_master,
        bra_code=branch_code,
    )


def get_user_master_context():
    return {
        "departments": (
            sys_dep_master.objects
            .all()
            .order_by("dep_code")
        ),
        "branches": (
            sys_bra_master.objects
            .all()
            .order_by("bra_code")
        ),
        "companies": (
            sys_cmp_master.objects
            .all()
            .order_by("cmp_code")
        ),
        "access_groups": (
            Group.objects
            .all()
            .order_by("name")
        ),
    }


def get_submitted_system_privileges(request):
    is_active = request.POST.get("is_active") == "on"

    if request.user.is_superuser:
        is_staff = request.POST.get("is_staff") == "on"
        is_superuser = request.POST.get("is_superuser") == "on"
    else:
        is_staff = False
        is_superuser = False

    if is_superuser:
        is_staff = True

    return is_active, is_staff, is_superuser


def sync_django_auth_user(
    user_obj,
    first_name,
    last_name,
    raw_password=None,
    old_login_id=None,
):
    login_id = (user_obj.usr_loginID or "").strip().lower()

    django_user = None

    if old_login_id:
        django_user = (
            User.objects
            .filter(username__iexact=old_login_id)
            .first()
        )

    if django_user is None:
        django_user = (
            User.objects
            .filter(username__iexact=login_id)
            .first()
        )

    if django_user is None:
        django_user = User(username=login_id)

    django_user.username = login_id
    django_user.email = (
        user_obj.usr_email or ""
    ).strip().lower()

    django_user.first_name = first_name
    django_user.last_name = last_name

    django_user.is_active = user_obj.usr_is_active
    django_user.is_staff = user_obj.usr_is_staff
    django_user.is_superuser = user_obj.usr_is_superuser

    auth_type = (user_obj.usr_auth or "").strip().upper()

    if auth_type == "LOCAL_DB":
        if raw_password:
            django_user.set_password(raw_password)

        elif not django_user.pk:
            django_user.set_unusable_password()

    else:
        django_user.set_unusable_password()

    django_user.save()

    django_user.groups.clear()

    if user_obj.usr_access_group:
        group = (
            Group.objects
            .filter(pk=user_obj.usr_access_group)
            .first()
        )

        if group:
            django_user.groups.add(group)

    return django_user


@login_required
@permission_required(
    "masters.view_sys_usr_system",
    raise_exception=True,
)
def user_master_list(request):
    users = (
        sys_usr_system.objects
        .select_related(
            "usr_dep_code",
            "usr_bra_code",
            "usr_company",
        )
        .all()
        .order_by("-id")
    )

    context = get_user_master_context()
    context["users"] = users

    return render(
        request,
        "masters/user_master_list.html",
        context,
    )


@login_required
@permission_required(
    "masters.add_sys_usr_system",
    raise_exception=True,
)
@transaction.atomic
def user_master_create(request):
    context = get_user_master_context()
    context["user_obj"] = None

    if request.method != "POST":
        return render(
            request,
            "masters/user_master_form.html",
            context,
        )

    auth_type = (
        request.POST.get("auth_type") or ""
    ).strip().upper()

    password = request.POST.get("password")

    is_active, is_staff, is_superuser = (
        get_submitted_system_privileges(request)
    )

    first_name = (
        request.POST.get("first_name") or ""
    ).strip()

    last_name = (
        request.POST.get("last_name") or ""
    ).strip()

    full_name = f"{first_name} {last_name}".strip()

    login_id = (
        request.POST.get("login_id") or ""
    ).strip().lower()

    email = (
        request.POST.get("email") or ""
    ).strip().lower()

    employee_id = (
        request.POST.get("employee_id") or ""
    ).strip()

    department_code = (
        request.POST.get("department") or ""
    ).strip()

    company_code = request.POST.get("company")
    branch_code = request.POST.get("branch")

    access_group_id = (
        request.POST.get("access_group") or ""
    ).strip()

    if not login_id:
        messages.error(request, "Login ID is required.")
        return redirect("user_master_list")

    if not email:
        messages.error(request, "Email is required.")
        return redirect("user_master_list")

    if not full_name:
        messages.error(request, "User name is required.")
        return redirect("user_master_list")

    if not department_code:
        messages.error(request, "Department is required.")
        return redirect("user_master_list")

    if auth_type not in ["SSO", "LOCAL_DB"]:
        messages.error(
            request,
            "Please select a valid authentication type.",
        )
        return redirect("user_master_list")

    if auth_type == "LOCAL_DB" and not password:
        messages.error(
            request,
            "Password is required for a Local DB user.",
        )
        return redirect("user_master_list")

    if not access_group_id and not is_superuser:
        messages.error(
            request,
            "Access Group is required for a non-superuser account.",
        )
        return redirect("user_master_list")

    if sys_usr_system.objects.filter(
        usr_loginID__iexact=login_id
    ).exists():
        messages.error(
            request,
            f"A user with Login ID {login_id} already exists.",
        )
        return redirect("user_master_list")

    if User.objects.filter(
        username__iexact=login_id
    ).exists():
        messages.error(
            request,
            f"A Django authentication user with Login ID "
            f"{login_id} already exists.",
        )
        return redirect("user_master_list")

    if sys_usr_system.objects.filter(
        usr_email__iexact=email
    ).exists():
        messages.error(
            request,
            f"A user with email {email} already exists.",
        )
        return redirect("user_master_list")

    usr_pno = employee_id or login_id

    if sys_usr_system.objects.filter(
        usr_pno=usr_pno
    ).exists():
        messages.error(
            request,
            f"A user with PNO/identifier {usr_pno} already exists.",
        )
        return redirect("user_master_list")

    department = get_object_or_404(
        sys_dep_master,
        dep_code=department_code,
    )

    company = get_optional_company(company_code)
    branch = get_optional_branch(branch_code)

    if access_group_id:
        access_group = Group.objects.filter(
            pk=access_group_id
        ).first()

        if not access_group:
            messages.error(
                request,
                "The selected Access Group does not exist.",
            )
            return redirect("user_master_list")

    user_obj = sys_usr_system(
        usr_pno=usr_pno,
        usr_name=full_name,
        usr_designation=(
            request.POST.get("designation") or ""
        ).strip(),
        usr_dep_code=department,
        usr_mobile=(
            request.POST.get("mobile") or ""
        ).strip() or None,
        usr_email=email,
        usr_phone=(
            request.POST.get("phone") or ""
        ).strip() or None,
        usr_loginID=login_id,
        usr_auth=auth_type,
        usr_access_group=access_group_id,
        usr_bra_code=branch,
        usr_company=company,
        usr_is_active=is_active,
        usr_is_staff=is_staff,
        usr_is_superuser=is_superuser,
    )

    if auth_type == "LOCAL_DB":
        user_obj.usr_password = make_password(password)
    else:
        user_obj.usr_password = None

    user_obj.save()

    sync_django_auth_user(
        user_obj=user_obj,
        first_name=first_name,
        last_name=last_name,
        raw_password=password,
    )

    messages.success(
        request,
        "User created successfully.",
    )

    return redirect("user_master_list")


@login_required
@permission_required(
    "masters.change_sys_usr_system",
    raise_exception=True,
)
@transaction.atomic
def user_master_update(request, pk):
    user_obj = get_object_or_404(
        sys_usr_system,
        pk=pk,
    )

    context = get_user_master_context()
    context["user_obj"] = user_obj

    if request.method != "POST":
        full_name = (user_obj.usr_name or "").strip()
        name_parts = full_name.split(" ", 1)

        user_obj.first_name_display = (
            name_parts[0] if name_parts else ""
        )

        user_obj.last_name_display = (
            name_parts[1]
            if len(name_parts) > 1
            else ""
        )

        return render(
            request,
            "masters/user_master_form.html",
            context,
        )

    old_login_id = user_obj.usr_loginID

    linked_django_user = (
        User.objects
        .filter(username__iexact=old_login_id)
        .first()
    )

    auth_type = (
        request.POST.get("auth_type") or ""
    ).strip().upper()

    password = request.POST.get("password")

    submitted_active = (
        request.POST.get("is_active") == "on"
    )

    if request.user.is_superuser:
        submitted_staff = (
            request.POST.get("is_staff") == "on"
        )
        submitted_superuser = (
            request.POST.get("is_superuser") == "on"
        )
    else:
        submitted_staff = user_obj.usr_is_staff
        submitted_superuser = user_obj.usr_is_superuser

    if submitted_superuser:
        submitted_staff = True

    editing_own_account = (
        linked_django_user is not None
        and linked_django_user.pk == request.user.pk
    )

    if editing_own_account:
        if not submitted_active:
            messages.error(
                request,
                "You cannot deactivate your own account.",
            )
            return redirect("user_master_list")

        if request.user.is_superuser and not submitted_superuser:
            messages.error(
                request,
                "You cannot remove your own superuser status.",
            )
            return redirect("user_master_list")

        if request.user.is_staff and not submitted_staff:
            messages.error(
                request,
                "You cannot remove your own staff status.",
            )
            return redirect("user_master_list")

    first_name = (
        request.POST.get("first_name") or ""
    ).strip()

    last_name = (
        request.POST.get("last_name") or ""
    ).strip()

    full_name = f"{first_name} {last_name}".strip()

    login_id = (
        request.POST.get("login_id") or ""
    ).strip().lower()

    email = (
        request.POST.get("email") or ""
    ).strip().lower()

    employee_id = (
        request.POST.get("employee_id") or ""
    ).strip()

    department_code = (
        request.POST.get("department") or ""
    ).strip()

    access_group_id = (
        request.POST.get("access_group") or ""
    ).strip()

    if not login_id:
        messages.error(request, "Login ID is required.")
        return redirect("user_master_list")

    if not email:
        messages.error(request, "Email is required.")
        return redirect("user_master_list")

    if not full_name:
        messages.error(request, "User name is required.")
        return redirect("user_master_list")

    if not department_code:
        messages.error(request, "Department is required.")
        return redirect("user_master_list")

    if auth_type not in ["SSO", "LOCAL_DB"]:
        messages.error(
            request,
            "Please select a valid authentication type.",
        )
        return redirect("user_master_list")

    if not access_group_id and not submitted_superuser:
        messages.error(
            request,
            "Access Group is required for a non-superuser account.",
        )
        return redirect("user_master_list")

    duplicate_login = (
        sys_usr_system.objects
        .filter(usr_loginID__iexact=login_id)
        .exclude(pk=user_obj.pk)
        .exists()
    )

    if duplicate_login:
        messages.error(
            request,
            f"A user with Login ID {login_id} already exists.",
        )
        return redirect("user_master_list")

    duplicate_auth_user = (
        User.objects
        .filter(username__iexact=login_id)
    )

    if linked_django_user:
        duplicate_auth_user = duplicate_auth_user.exclude(
            pk=linked_django_user.pk
        )

    if duplicate_auth_user.exists():
        messages.error(
            request,
            f"A Django authentication user with Login ID "
            f"{login_id} already exists.",
        )
        return redirect("user_master_list")

    duplicate_email = (
        sys_usr_system.objects
        .filter(usr_email__iexact=email)
        .exclude(pk=user_obj.pk)
        .exists()
    )

    if duplicate_email:
        messages.error(
            request,
            f"A user with email {email} already exists.",
        )
        return redirect("user_master_list")

    usr_pno = employee_id or login_id

    duplicate_pno = (
        sys_usr_system.objects
        .filter(usr_pno=usr_pno)
        .exclude(pk=user_obj.pk)
        .exists()
    )

    if duplicate_pno:
        messages.error(
            request,
            f"A user with PNO/identifier {usr_pno} already exists.",
        )
        return redirect("user_master_list")

    department = get_object_or_404(
        sys_dep_master,
        dep_code=department_code,
    )

    company = get_optional_company(
        request.POST.get("company")
    )

    branch = get_optional_branch(
        request.POST.get("branch")
    )

    if access_group_id:
        access_group = Group.objects.filter(
            pk=access_group_id
        ).first()

        if not access_group:
            messages.error(
                request,
                "The selected Access Group does not exist.",
            )
            return redirect("user_master_list")

    user_obj.usr_pno = usr_pno
    user_obj.usr_name = full_name
    user_obj.usr_designation = (
        request.POST.get("designation") or ""
    ).strip()
    user_obj.usr_dep_code = department
    user_obj.usr_mobile = (
        request.POST.get("mobile") or ""
    ).strip() or None
    user_obj.usr_email = email
    user_obj.usr_phone = (
        request.POST.get("phone") or ""
    ).strip() or None
    user_obj.usr_loginID = login_id
    user_obj.usr_auth = auth_type
    user_obj.usr_access_group = access_group_id
    user_obj.usr_bra_code = branch
    user_obj.usr_company = company

    user_obj.usr_is_active = submitted_active
    user_obj.usr_is_staff = submitted_staff
    user_obj.usr_is_superuser = submitted_superuser

    if auth_type == "LOCAL_DB":
        if password:
            user_obj.usr_password = make_password(password)
    else:
        user_obj.usr_password = None

    user_obj.save()

    sync_django_auth_user(
        user_obj=user_obj,
        first_name=first_name,
        last_name=last_name,
        raw_password=password,
        old_login_id=old_login_id,
    )

    messages.success(
        request,
        "User updated successfully.",
    )

    return redirect("user_master_list")


@login_required
@permission_required(
    "masters.delete_sys_usr_system",
    raise_exception=True,
)
@transaction.atomic
def user_master_delete(request, pk):
    user_obj = get_object_or_404(
        sys_usr_system,
        pk=pk,
    )

    if request.method != "POST":
        return redirect("user_master_list")

    linked_django_user = (
        User.objects
        .filter(username__iexact=user_obj.usr_loginID)
        .first()
    )

    if (
        linked_django_user
        and linked_django_user.pk == request.user.pk
    ):
        messages.error(
            request,
            "You cannot delete your own logged-in account.",
        )
        return redirect("user_master_list")

    if (
        user_obj.usr_is_superuser
        and not request.user.is_superuser
    ):
        messages.error(
            request,
            "Only a superuser can delete another superuser.",
        )
        return redirect("user_master_list")

    if linked_django_user:
        linked_django_user.delete()

    user_obj.delete()

    messages.success(
        request,
        "User deleted successfully.",
    )

    return redirect("user_master_list")


@login_required
@permission_required("auth.view_group", raise_exception=True)
def access_group_list(request):
    groups = (
        Group.objects
        .prefetch_related("permissions", "user_set")
        .all()
        .order_by("name")
    )

    permissions = (
        Permission.objects
        .select_related("content_type")
        .exclude(content_type__app_label__in=[
            "admin",
            "contenttypes",
            "sessions",
            "sites",
            "account",
            "socialaccount",
        ])
        .order_by(
            "content_type__app_label",
            "content_type__model",
            "codename",
        )
    )

    grouped_permissions = defaultdict(list)

    for permission in permissions:
        app_label = permission.content_type.app_label
        model_name = permission.content_type.model

        group_key = f"{app_label}.{model_name}"

        grouped_permissions[group_key].append(permission)

    grouped_permissions = dict(grouped_permissions)

    return render(
        request,
        "masters/access_group_list.html",
        {
            "groups": groups,
            "grouped_permissions": grouped_permissions,
        },
    )


@login_required
@permission_required("auth.add_group", raise_exception=True)
def access_group_create(request):
    if request.method == "POST":
        group_name = request.POST.get("group_name", "").strip()

        if Group.objects.filter(name=group_name).exists():
            messages.error(request, "Access group already exists.")
            return redirect("access_group_list")

        group = Group.objects.create(name=group_name)

        permission_ids = request.POST.getlist("permissions")

        group.permissions.set(
            Permission.objects.filter(id__in=permission_ids)
        )

        messages.success(request, "Access group created successfully.")

    return redirect("access_group_list")


@login_required
@permission_required("auth.change_group", raise_exception=True)
def access_group_update(request, pk):
    group = get_object_or_404(Group, pk=pk)

    if request.method == "POST":
        group.name = request.POST.get("group_name").strip()
        group.save()

        permission_ids = request.POST.getlist("permissions")

        group.permissions.set(
            Permission.objects.filter(id__in=permission_ids)
        )

        messages.success(request, "Access group updated successfully.")

    return redirect("access_group_list")


@login_required
@permission_required("auth.delete_group", raise_exception=True)
def access_group_delete(request, pk):
    group = get_object_or_404(Group, pk=pk)

    if request.method == "POST":
        group.delete()

        messages.success(
            request,
            "Access group deleted successfully."
        )

    return redirect("access_group_list")


@login_required
@permission_required("masters.change_sys_usr_system", raise_exception=True)
def database_settings_panel(request):
    env_path = os.path.join(settings.BASE_DIR, '.env')
    
    current_settings = {
        'DB_NAME': '', 
        'DB_USER': '', 
        'DB_PASSWORD': '', 
        'DB_HOST': '127.0.0.1', 
        'DB_PORT': '3306'
    }
    
    if os.path.exists(env_path):
        try:
            with open(env_path, 'r') as file:
                for line in file:
                    cleaned_line = line.strip()
                    if '=' in cleaned_line and not cleaned_line.startswith('#'):
                        key, value = cleaned_line.split('=', 1)
                        if key in current_settings:
                            current_settings[key] = value
        except Exception as e:
            messages.error(request, f"Error reading system config: {str(e)}")

    if request.method == 'POST':
        db_name = request.POST.get('db_name', '').strip()
        db_user = request.POST.get('db_user', '').strip()
        db_password = request.POST.get('db_password', '').strip()
        db_host = request.POST.get('db_host', '').strip() or '127.0.0.1'
        db_port = request.POST.get('db_port', '').strip() or '3306'

        try:
            test_conn = MySQLdb.connect(
                host=db_host,
                user=db_user,
                passwd=db_password,
                db=db_name,
                port=int(db_port),
                connect_timeout=5
            )
            test_conn.close()
            
        except Exception as db_error:
            messages.error(
                request, 
                f"Database Connection Refused! The website would crash if saved. "
                f"Verify your MySQL credentials or service state. Error: {str(db_error)}"
            )
            submitted_settings = {
                'DB_NAME': db_name, 'DB_USER': db_user, 
                'DB_PASSWORD': db_password, 'DB_HOST': db_host, 'DB_PORT': db_port
            }
            return render(request, 'masters/database_settings.html', {'current_settings': submitted_settings})

        env_content = f"""# Highnoon database settings
DB_NAME={db_name}
DB_USER={db_user}
DB_PASSWORD={db_password}
DB_HOST={db_host}
DB_PORT={db_port}
DEBUG=False
"""
        try:
            with open(env_path, 'w') as file:
                file.write(env_content)
            
            messages.success(request, "Database credentials verified and saved successfully!")
            
            wsgi_path = os.path.join(settings.BASE_DIR, 'highnoon_vms', 'wsgi.py')
            if os.path.exists(wsgi_path):
                os.utime(wsgi_path, None)
            
            return redirect('database_settings_panel')
            
        except Exception as e:
            messages.error(request, f"File system write failure: {str(e)}")

    return render(request, 'masters/database_settings.html', {'current_settings': current_settings})